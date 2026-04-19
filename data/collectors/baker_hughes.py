"""
Baker Hughes North America Rig Count collector.

Downloads the weekly North America rig count Excel workbook from Baker Hughes'
public website and writes US and Canada series into the ``macro`` ArcticDB
library via ``store.write_series()``.

Series collected (weekly, keyed on Friday report date):
  - BHI_US_TOTAL_RIGS  : US total rig count
  - BHI_US_OIL_RIGS    : US oil-directed rig count
  - BHI_US_GAS_RIGS    : US gas-directed rig count
  - BHI_CANADA_RIGS    : Canada total rig count

ArcticDB key: as above (e.g. ``BHI_US_TOTAL_RIGS``)
Columns: ``count`` (integer rig count)
Index: UTC DatetimeIndex (weekly, Friday release date)

Incremental: reads last stored date; appends only new rows on subsequent runs.

Run via: ``python -m data.collectors.baker_hughes``
"""

import logging
from datetime import timedelta
from io import BytesIO

import pandas as pd
import requests  # type: ignore[import-untyped]

from data.store import get_store

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Baker Hughes public Excel download URL for North America rig count
# Updated weekly; no API key required
BHI_NA_EXCEL_URL = (
    "https://rigcount.bakerhughes.com/static-files/16a70c3d-0e8c-4ec1-afa4-f7ebd11c3120"
)

# Candidate sheet names for North America summary data (BHI renames occasionally)
_NA_SHEET_CANDIDATES = [
    "NAM Summary",
    "North America Rig Count",
    "NA Rig Count",
    "North America",
    "NA",
]

# ArcticDB keys for each series
SERIES_KEYS: list[str] = [
    "BHI_US_TOTAL_RIGS",
    "BHI_US_OIL_RIGS",
    "BHI_US_GAS_RIGS",
    "BHI_CANADA_RIGS",
]

# Baker Hughes public Excel download URL for Worldwide rig count (monthly)
# Updated monthly; no API key required
BHI_WW_EXCEL_URL = (
    "https://rigcount.bakerhughes.com/static-files/93ea6fff-3d5c-4747-84e8-4a59e8ddf827"
)

# Candidate sheet names for worldwide summary data
_WW_SHEET_CANDIDATES = [
    "WW Summary",
    "Worldwide Summary",
    "Worldwide Rig Count",
    "Summary",
]

# ArcticDB keys for worldwide series (monthly averages)
WW_SERIES_KEYS: list[str] = [
    "BHI_WW_TOTAL_RIGS",
    "BHI_WW_INTL_RIGS",
    "BHI_WW_NA_RIGS",
    "BHI_WW_US_RIGS",
    "BHI_WW_CANADA_RIGS",
    "BHI_WW_LATAM_RIGS",
    "BHI_WW_EUROPE_RIGS",
    "BHI_WW_AFRICA_RIGS",
    "BHI_WW_MIDEAST_RIGS",
    "BHI_WW_APAC_RIGS",
]


# ---------------------------------------------------------------------------
# Download + parse
# ---------------------------------------------------------------------------


def _download_excel(url: str) -> bytes:
    """Download Baker Hughes Excel workbook bytes from *url*.

    Args:
        url: Direct URL to the Excel (.xlsx) file.

    Returns:
        Raw bytes of the Excel file.

    Raises:
        requests.HTTPError: On non-2xx response.
    """
    headers = {
        "User-Agent": ("Mozilla/5.0 (compatible; QuantWorkstation/1.0; +https://github.com/)")
    }
    response = requests.get(url, headers=headers, timeout=60)
    response.raise_for_status()
    return bytes(response.content)


def _parse_na_rig_count(excel_bytes: bytes) -> dict[str, pd.DataFrame]:
    """Parse North America rig count from Excel workbook bytes.

    BHI changed their format in early 2026. The new format (NAM Summary sheet)
    is a weekly snapshot, not a time series. Each file contains one week's
    counts. The date is in cell B4, and counts are found by matching row labels.

    Row labels we extract (col 0 = label, col 1 = this week's count):
      - "United States Total" → BHI_US_TOTAL_RIGS
      - "Oil"                 → BHI_US_OIL_RIGS   (U.S. Breakout section)
      - "Gas"                 → BHI_US_GAS_RIGS   (U.S. Breakout section)
      - "Canada"              → BHI_CANADA_RIGS

    Args:
        excel_bytes: Raw bytes of the Baker Hughes NA Excel workbook.

    Returns:
        Dict mapping ArcticDB key → single-row DataFrame with UTC
        DatetimeIndex and column ``count`` (Int64).

    Raises:
        ValueError: If the sheet or expected rows are not found.
    """
    import openpyxl as _openpyxl

    wb = _openpyxl.load_workbook(BytesIO(excel_bytes), read_only=True)
    available = wb.sheetnames
    wb.close()

    sheet_name = next((c for c in _NA_SHEET_CANDIDATES if c in available), None)
    if sheet_name is None:
        raise ValueError(
            f"Cannot find NA rig count sheet. Available: {available}. "
            f"Candidates tried: {_NA_SHEET_CANDIDATES}"
        )
    log.info("Using sheet '%s'", sheet_name)

    df_raw = pd.read_excel(
        BytesIO(excel_bytes),
        sheet_name=sheet_name,
        engine="openpyxl",
        header=None,
    )

    # Date is in row 3 (0-indexed), col 3.
    # BHI uses DD/MM/YYYY strings in newer files (e.g. "10/04/2026" = Apr 10).
    # Use dayfirst=True to avoid misparse as October.
    report_date = pd.to_datetime(df_raw.iloc[3, 3], dayfirst=True, errors="coerce", utc=False)
    if pd.isna(report_date):
        raise ValueError(f"Cannot parse report date from row 3 col 3: {df_raw.iloc[3, 3]!r}")
    report_date = report_date.tz_localize("UTC")
    log.info("Report date: %s", report_date.date())

    # Build label → this-week-count map from col 1 (label) / col 3 (this week)
    label_map: dict[str, int] = {}
    for _, row in df_raw.iterrows():
        label = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
        value = row.iloc[3]
        if label and label not in label_map and pd.notna(value):
            try:
                label_map[label] = int(value)
            except (ValueError, TypeError):
                pass

    def _find_count(*candidates: str) -> int:
        for c in candidates:
            if c in label_map:
                return label_map[c]
        raise ValueError(
            f"Cannot find row matching {candidates}. Available labels: {list(label_map.keys())}"
        )

    us_total = _find_count("United States Total", "U.S. Total", "US Total")
    us_oil = _find_count("Oil")
    us_gas = _find_count("Gas")
    canada = _find_count("Canada")

    def _make_df(count: int) -> pd.DataFrame:
        return pd.DataFrame({"count": [count]}, index=[report_date])

    return {
        "BHI_US_TOTAL_RIGS": _make_df(us_total),
        "BHI_US_OIL_RIGS": _make_df(us_oil),
        "BHI_US_GAS_RIGS": _make_df(us_gas),
        "BHI_CANADA_RIGS": _make_df(canada),
    }


# ---------------------------------------------------------------------------
# Public collect function
# ---------------------------------------------------------------------------


def collect(url: str = BHI_NA_EXCEL_URL) -> None:
    """Download Baker Hughes NA rig count and write to ArcticDB ``macro`` library.

    Incremental: reads the last stored date for each series and appends only
    rows newer than that date.  Safe to run repeatedly — no duplicates.

    Args:
        url: Direct URL to the Baker Hughes NA rig count Excel file.
             Defaults to ``BHI_NA_EXCEL_URL``.
    """
    log.info("Downloading Baker Hughes NA rig count from %s", url)
    excel_bytes = _download_excel(url)
    log.info("Downloaded %d bytes; parsing…", len(excel_bytes))

    try:
        series_map = _parse_na_rig_count(excel_bytes)
    except (ValueError, IndexError, KeyError) as exc:
        log.warning("Failed to parse BHI NA rig count Excel: %s — skipping", exc)
        return
    store = get_store()

    for arc_key, df_new in series_map.items():
        if df_new.empty:
            log.warning("Series %s: parsed data is empty — skipping", arc_key)
            continue

        # Incremental: find last stored date and filter to new-only rows
        try:
            existing = store.read_series("macro", arc_key)
            if not existing.empty:
                last_date = existing.index.max()
                cutoff = last_date + timedelta(days=1)
                df_new = df_new[df_new.index >= cutoff]
                log.info(
                    "Series %s: last stored %s — %d new rows to append",
                    arc_key,
                    last_date.date(),
                    len(df_new),
                )
        except Exception:
            log.info("Series %s: no existing data — writing full history", arc_key)

        if df_new.empty:
            log.info("Series %s: no new data to write", arc_key)
            continue

        store.write_series("macro", arc_key, df_new)
        log.info("Wrote %d rows to macro/%s", len(df_new), arc_key)

    log.info("Baker Hughes collection complete")


# ---------------------------------------------------------------------------
# Worldwide parser + collect
# ---------------------------------------------------------------------------


def _parse_ww_rig_count(excel_bytes: bytes) -> dict[str, pd.DataFrame]:
    """Parse Worldwide rig count from Excel workbook bytes.

    The WW Summary sheet is a monthly snapshot. Layout (0-indexed):
      - Row 1: dates at col 4 (current month), col 10 (prev month), col 16 (year ago)
      - Row 2: column headers (Land, Offshore, Total, ...)
      - Label col: 1 — Total (current month) col: 5

    Rows extracted:
      row 4  Latin America   → BHI_WW_LATAM_RIGS
      row 5  Europe          → BHI_WW_EUROPE_RIGS
      row 6  Africa          → BHI_WW_AFRICA_RIGS
      row 7  Middle East     → BHI_WW_MIDEAST_RIGS
      row 8  Asia-Pacific    → BHI_WW_APAC_RIGS
      row 9  International   → BHI_WW_INTL_RIGS
      row 11 United States   → BHI_WW_US_RIGS
      row 12 Canada          → BHI_WW_CANADA_RIGS
      row 13 North America   → BHI_WW_NA_RIGS
      row 15 Worldwide       → BHI_WW_TOTAL_RIGS

    Values are monthly averages (floats), stored as float64.

    Args:
        excel_bytes: Raw bytes of the Baker Hughes WW Excel workbook.

    Returns:
        Dict mapping ArcticDB key → single-row DataFrame with UTC
        DatetimeIndex and column ``count`` (float64).

    Raises:
        ValueError: If the sheet or expected rows are not found.
    """
    import openpyxl as _openpyxl

    wb = _openpyxl.load_workbook(BytesIO(excel_bytes), read_only=True)
    available = wb.sheetnames
    wb.close()

    sheet_name = next((c for c in _WW_SHEET_CANDIDATES if c in available), None)
    if sheet_name is None:
        raise ValueError(
            f"Cannot find WW rig count sheet. Available: {available}. "
            f"Candidates tried: {_WW_SHEET_CANDIDATES}"
        )
    log.info("Using WW sheet '%s'", sheet_name)

    df_raw = pd.read_excel(
        BytesIO(excel_bytes),
        sheet_name=sheet_name,
        engine="openpyxl",
        header=None,
    )

    # Date is in row 1, col 4 (current month)
    report_date = pd.to_datetime(df_raw.iloc[1, 4], errors="coerce", utc=False)
    if pd.isna(report_date):
        raise ValueError(f"Cannot parse report date from row 1 col 4: {df_raw.iloc[1, 4]!r}")
    report_date = report_date.tz_localize("UTC")
    log.info("WW report date: %s", report_date.date())

    # Build label → total (col 5) map; first occurrence wins
    label_map: dict[str, float] = {}
    for _, row in df_raw.iterrows():
        label = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
        value = row.iloc[5]
        if label and label not in label_map and pd.notna(value):
            try:
                label_map[label] = float(value)
            except (ValueError, TypeError):
                pass

    def _find_count(*candidates: str) -> float:
        for c in candidates:
            if c in label_map:
                return label_map[c]
        raise ValueError(
            f"Cannot find row matching {candidates}. Available labels: {list(label_map.keys())}"
        )

    def _make_df(count: float) -> pd.DataFrame:
        return pd.DataFrame({"count": [count]}, index=[report_date])

    return {
        "BHI_WW_TOTAL_RIGS": _make_df(_find_count("Worldwide")),
        "BHI_WW_INTL_RIGS": _make_df(_find_count("International")),
        "BHI_WW_NA_RIGS": _make_df(_find_count("North America")),
        "BHI_WW_US_RIGS": _make_df(_find_count("United States")),
        "BHI_WW_CANADA_RIGS": _make_df(_find_count("Canada")),
        "BHI_WW_LATAM_RIGS": _make_df(_find_count("Latin America")),
        "BHI_WW_EUROPE_RIGS": _make_df(_find_count("Europe")),
        "BHI_WW_AFRICA_RIGS": _make_df(_find_count("Africa")),
        "BHI_WW_MIDEAST_RIGS": _make_df(_find_count("Middle East")),
        "BHI_WW_APAC_RIGS": _make_df(_find_count("Asia-Pacific", "Asia Pacific")),
    }


def collect_worldwide(url: str = BHI_WW_EXCEL_URL) -> None:
    """Download Baker Hughes WW rig count and write to ArcticDB ``macro`` library.

    Incremental: reads the last stored date for each series and appends only
    rows newer than that date.  Safe to run repeatedly — no duplicates.

    Args:
        url: Direct URL to the Baker Hughes WW rig count Excel file.
             Defaults to ``BHI_WW_EXCEL_URL``.
    """
    log.info("Downloading Baker Hughes WW rig count from %s", url)
    excel_bytes = _download_excel(url)
    log.info("Downloaded %d bytes; parsing…", len(excel_bytes))

    series_map = _parse_ww_rig_count(excel_bytes)
    store = get_store()

    for arc_key, df_new in series_map.items():
        if df_new.empty:
            log.warning("Series %s: parsed data is empty — skipping", arc_key)
            continue

        try:
            existing = store.read_series("macro", arc_key)
            if not existing.empty:
                last_date = existing.index.max()
                cutoff = last_date + timedelta(days=1)
                df_new = df_new[df_new.index >= cutoff]
                log.info(
                    "Series %s: last stored %s — %d new rows to append",
                    arc_key,
                    last_date.date(),
                    len(df_new),
                )
        except Exception:
            log.info("Series %s: no existing data — writing full history", arc_key)

        if df_new.empty:
            log.info("Series %s: no new data to write", arc_key)
            continue

        store.write_series("macro", arc_key, df_new)
        log.info("Wrote %d rows to macro/%s", len(df_new), arc_key)

    log.info("Baker Hughes WW collection complete")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    collect()
    collect_worldwide()
